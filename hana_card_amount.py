import openpyxl
import re


def get_hana_amount() :
    # Open the workbook
    workbook = openpyxl.load_workbook("input\\input_hana.xlsx")

    # Select the active sheet
    sheet = workbook.active

    final_amount = None

    # Iterate through the rows
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell is a string and contains '결제금액'
            if isinstance(cell.value, str) and '결제금액' in cell.value:
                # Get the column index of the found cell
                row_index = cell.row - 1  # openpyxl uses 1-based indexing
                col_index = cell.column - 1  # openpyxl uses 1-based indexing

                # Extract the amount from the cell to the right of the merged cell
                if row_index + 1 < sheet.max_row:
                    amount_cell = sheet.cell(row=row_index + 2, column=col_index + 1) # openpyxl uses 1-based indexing
                    if amount_cell.value is not None:
                        final_amount = amount_cell.value
                        # Try to convert to float
                        if isinstance(final_amount, str):
                            try:
                                final_amount = float(final_amount.replace(',', ''))
                            except ValueError:
                                final_amount = None  # If conversion fails, set to None
                        elif isinstance(final_amount, (int, float)):
                            final_amount = float(final_amount)  # Ensure it's a float
                        break  # Stop searching for cells in this row
        if final_amount is not None:
            break  # Stop searching for rows

    return final_amount

if __name__ == '__main__':
    final_amount = get_hana_amount()

    # Print the final amount if found
    if final_amount is not None:
        print("Final amount:", final_amount)
    else:
        print("Final amount not found in the sheet.")