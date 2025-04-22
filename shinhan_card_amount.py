import openpyxl
import re

# Open the workbook
workbook = openpyxl.load_workbook("input\\input_shinhan.xlsx")

# Select the active sheet
sheet = workbook.active

final_amount = None

# Function to check if a cell is merged and get its width
def get_merged_cell_width(sheet, cell):
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return merged_range.max_col - merged_range.min_col + 1
    return 1  # Default width if not merged

# Iterate through the rows
for row in sheet.iter_rows():
    for cell in row:
        # Check if the cell is a string and contains '입금하실 금액'
        if isinstance(cell.value, str) and '입금하실 금액' in cell.value:
            # Get the column index of the found cell
            col_index = cell.column - 1  # openpyxl uses 1-based indexing

            # Get the width of the merged cell (if any)
            merged_cell_width = get_merged_cell_width(sheet, cell)

            # Extract the amount from the cell to the right of the merged cell
            amount_col_index = col_index + merged_cell_width
            if amount_col_index < sheet.max_column:
                amount_cell = row[amount_col_index]
                if amount_cell.value is not None:
                    final_amount = amount_cell.value
                    # Try to convert to float
                    if isinstance(final_amount, str):
                        try:
                            final_amount = float(final_amount.replace(',', ''))
                        except ValueError:
                            final_amount = None  # If conversion fails, set to None
                    elif isinstance(final_amount, (int, float)):
                        final_amount = float(final_amount)  # Ensure it's a float
                    break  # Stop searching for cells in this row
    if final_amount is not None:
        break  # Stop searching for rows

# Print the final amount if found
if final_amount is not None:
    print("Final amount:", final_amount)
else:
    print("Final amount not found in the sheet.")