import openpyxl
import re

# Open the workbook
workbook = openpyxl.load_workbook("input\\input_samsung.xlsx")

# Select the active sheet
sheet = workbook['일시불']

final_amount = None

# Iterate through the rows
for row in sheet.iter_rows():
    for cell in row:
        # Check if the cell is a string and contains '일시불합계'
        if isinstance(cell.value, str) and '일시불합계' in cell.value:
            # Get the column index of the found cell
            col_index = cell.column - 1  # openpyxl uses 1-based indexing

            # Get the width of the merged cell (if any)
            total_amount = 7

            # Extract the amount from the cell to the right of the merged cell
            amount_col_index = col_index + total_amount
            if amount_col_index < sheet.max_column:
                amount_cell = row[amount_col_index]
                if amount_cell.value is not None:
                    final_amount = amount_cell.value
                    break  # Stop searching for cells in this row
    if final_amount is not None:
        break  # Stop searching for rows

# Print the final amount if found
if final_amount is not None:
    print("Final amount:", final_amount)
else:
    print("Final amount not found in the sheet.")

# sheet = workbook['할부']

# final_amount1 = None

# # Iterate through the rows
# for row in sheet.iter_rows():
#     for cell in row:
#         # Check if the cell is a string and contains '할부합계'
#         if isinstance(cell.value, str) and '할부합계' in cell.value:
#             # Get the column index of the found cell
#             col_index = cell.column - 1  # openpyxl uses 1-based indexing

#             # Get the width of the merged cell (if any)
#             total_amount = 7

#             # Extract the amount from the cell to the right of the merged cell
#             amount_col_index = col_index + total_amount
#             if amount_col_index < sheet.max_column:
#                 amount_cell = row[amount_col_index]
#                 if amount_cell.value is not None:
#                     final_amount1 = amount_cell.value
#                     break  # Stop searching for cells in this row
#     if final_amount1 is not None:
#         break  # Stop searching for rows

# # Print the final amount if found
# if final_amount1 is not None and final_amount is not None:
#     print("Final amount:", final_amount)
#     print("Final amount:", final_amount1)

#     final_total_amount = final_amount + final_amount1

#     print("Final total amount:", final_total_amount)
# else:
#     print("Final amount not found in the sheet.")